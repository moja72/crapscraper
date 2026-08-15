from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


PASTA = Path("data/comparacao/entrada")
ARQUIVOS = sorted(PASTA.glob("*.xlsx"))

if not ARQUIVOS:
    raise SystemExit(f"Nenhuma planilha .xlsx encontrada em: {PASTA.resolve()}")


for arquivo in ARQUIVOS:
    print("=" * 110)
    print(f"ARQUIVO: {arquivo.name}")
    print(f"CAMINHO: {arquivo.resolve()}")

    workbook = load_workbook(
        arquivo,
        read_only=True,
        data_only=True,
    )

    print(f"ABAS: {', '.join(workbook.sheetnames)}")

    for nome_aba in workbook.sheetnames:
        planilha = workbook[nome_aba]

        print()
        print(f"ABA: {nome_aba}")
        print(
            f"LINHAS APROXIMADAS: {planilha.max_row} | "
            f"COLUNAS APROXIMADAS: {planilha.max_column}"
        )
        print("PRIMEIRAS 5 LINHAS PREENCHIDAS:")

        linhas_exibidas = 0

        for numero_linha, linha in enumerate(
            planilha.iter_rows(values_only=True),
            start=1,
        ):
            celulas = []

            for numero_coluna, valor in enumerate(linha, start=1):
                if valor is None or str(valor).strip() == "":
                    continue

                coluna = get_column_letter(numero_coluna)
                tipo = type(valor).__name__
                celulas.append(f"{coluna}={valor!r} <{tipo}>")

            if not celulas:
                continue

            print(f"LINHA {numero_linha}:")
            print(" | ".join(celulas))
            linhas_exibidas += 1

            if linhas_exibidas >= 5:
                break

        if linhas_exibidas == 0:
            print("A aba não possui linhas preenchidas.")

    workbook.close()
    print()