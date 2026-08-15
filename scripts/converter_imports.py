from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import load_workbook


IMPORTS_DIR = Path("data") / "imports"


def converter_xlsx_para_csv(arquivo_xlsx: Path) -> Path:
    workbook = load_workbook(
        arquivo_xlsx,
        read_only=True,
        data_only=True,
    )

    try:
        planilha = workbook[workbook.sheetnames[0]]
        arquivo_csv = arquivo_xlsx.with_suffix(".csv")

        with arquivo_csv.open(
            "w",
            encoding="utf-8-sig",
            newline="",
        ) as destino:
            writer = csv.writer(
                destino,
                delimiter=",",
                quoting=csv.QUOTE_MINIMAL,
            )

            for linha in planilha.iter_rows(values_only=True):
                writer.writerow(
                    [
                        "" if valor is None else valor
                        for valor in linha
                    ]
                )

        return arquivo_csv

    finally:
        workbook.close()


def main() -> None:
    IMPORTS_DIR.mkdir(parents=True, exist_ok=True)

    arquivos_xlsx = sorted(IMPORTS_DIR.glob("*.xlsx"))

    if not arquivos_xlsx:
        print("Nenhum arquivo XLSX encontrado.")
        return

    print(f"Pasta: {IMPORTS_DIR.resolve()}")
    print()

    for arquivo_xlsx in arquivos_xlsx:
        arquivo_csv = converter_xlsx_para_csv(arquivo_xlsx)

        print(
            f"OK: {arquivo_xlsx.name} "
            f"-> {arquivo_csv.name}"
        )

    print()
    print("Conversão concluída.")


if __name__ == "__main__":
    main()